"""
مسارات المالية - Finance Routes
إدارة خطط العلاج والمدفوعات والأقساط
"""

import io
import importlib
import sqlite3
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    openpyxl = None
    Font = Alignment = PatternFill = None
from flask import Blueprint, render_template, request, jsonify, make_response
from routes.auth import login_required
from services.finance_service import (
    get_all_treatment_plans, create_treatment_plan,
    add_payment, get_payments_for_patient, get_treatment_plan_by_id,
    get_all_payments
)
from db.database import get_db

finance_bp = Blueprint('finance', __name__, url_prefix='/finance')


def _ensure_finance_schema(db):
    """Ensure finance-related columns exist for legacy databases."""
    plan_cols = {row['name'] for row in db.execute("PRAGMA table_info(treatment_plans)").fetchall()}
    if 'initial_payment' not in plan_cols:
        db.execute("ALTER TABLE treatment_plans ADD COLUMN initial_payment REAL NOT NULL DEFAULT 0")
    if 'installments_count' not in plan_cols:
        db.execute("ALTER TABLE treatment_plans ADD COLUMN installments_count INTEGER NOT NULL DEFAULT 1")
    if 'amount_paid' not in plan_cols:
        db.execute("ALTER TABLE treatment_plans ADD COLUMN amount_paid REAL NOT NULL DEFAULT 0")
    if 'status' not in plan_cols:
        db.execute("ALTER TABLE treatment_plans ADD COLUMN status TEXT NOT NULL DEFAULT 'active'")

    pay_cols = {row['name'] for row in db.execute("PRAGMA table_info(payments)").fetchall()}
    if 'treatment_plan_id' not in pay_cols:
        db.execute("ALTER TABLE payments ADD COLUMN treatment_plan_id INTEGER")

    db.commit()


@finance_bp.route('/')
@login_required
def index():
    """صفحة المالية"""
    db = get_db()
    plans = db.execute('''
        SELECT tp.*, p.full_name as patient_name
        FROM treatment_plans tp
        JOIN patients p ON tp.patient_id = p.id
        ORDER BY tp.created_at DESC
    ''').fetchall()

    payments = db.execute('''
        SELECT pay.*, p.full_name as patient_name,
               ROW_NUMBER() OVER (
                   PARTITION BY pay.patient_id
                   ORDER BY pay.payment_date, pay.id
               ) as payment_num
        FROM payments pay
        JOIN patients p ON pay.patient_id = p.id
        ORDER BY pay.payment_date, pay.id
        LIMIT 50
    ''').fetchall()

    patients = db.execute('SELECT id, full_name FROM patients ORDER BY id').fetchall()

    # ─── ملخص مالي ────────────────────────────────────────────────
    total_income    = db.execute('SELECT COALESCE(SUM(amount), 0) as t FROM payments').fetchone()['t']
    total_remaining = db.execute(
        'SELECT COALESCE(SUM(total_cost - amount_paid), 0) as t FROM treatment_plans WHERE status="active"'
    ).fetchone()['t']

    return render_template('finance.html',
        plans=plans,
        payments=payments,
        patients=patients,
        total_income=total_income,
        total_remaining=total_remaining
    )


@finance_bp.route('/plan/add', methods=['POST'])
@login_required
def add_plan():
    """إضافة خطة علاج جديدة"""
    data = {
        'patient_id':         request.form.get('patient_id'),
        'description':        request.form.get('description', '').strip(),
        'total_cost':         request.form.get('total_cost', 0),
        'initial_payment':    request.form.get('initial_payment', 0),
        'installments_count': request.form.get('installments_count', 1),
    }

    if not data['patient_id'] or not data['description']:
        return jsonify({'success': False, 'message': 'المريض والوصف مطلوبان'}), 400

    try:
        data['total_cost']         = float(data['total_cost'])
        data['initial_payment']    = float(data['initial_payment'])
        data['installments_count'] = int(data['installments_count'])
    except ValueError:
        return jsonify({'success': False, 'message': 'قيم غير صالحة'}), 400

    plan_id = create_treatment_plan(data)
    return jsonify({'success': True, 'message': 'تم إنشاء خطة العلاج', 'id': plan_id})


@finance_bp.route('/payment/add', methods=['POST'])
@login_required
def add_payment_route():
    """تسجيل دفعة جديدة"""
    data = {
        'patient_id':        request.form.get('patient_id'),
        'treatment_plan_id': request.form.get('treatment_plan_id') or None,
        'amount':            request.form.get('amount', 0),
        'payment_date':      request.form.get('payment_date', ''),
        'payment_method':    request.form.get('payment_method', 'cash'),
        'notes':             request.form.get('notes', '').strip(),
    }

    if not data['patient_id'] or not data['amount']:
        return jsonify({'success': False, 'message': 'المريض والمبلغ مطلوبان'}), 400

    try:
        data['amount'] = float(data['amount'])
    except ValueError:
        return jsonify({'success': False, 'message': 'مبلغ غير صالح'}), 400

    payment_id = add_payment(data)
    return jsonify({'success': True, 'message': 'تم تسجيل الدفعة', 'id': payment_id})


@finance_bp.route('/plan/<int:plan_id>')
@login_required
def plan_detail(plan_id):
    """تفاصيل خطة علاج (JSON)"""
    plan = get_treatment_plan_by_id(plan_id)
    if not plan:
        return jsonify({'success': False, 'message': 'الخطة غير موجودة'}), 404
    return jsonify({'success': True, 'plan': dict(plan)})


@finance_bp.route('/plan/<int:plan_id>/delete', methods=['POST'])
@login_required
def delete_plan(plan_id):
    """حذف خطة علاج وكل مدفوعاتها المرتبطة"""
    db = get_db()
    _ensure_finance_schema(db)
    plan = db.execute('SELECT id FROM treatment_plans WHERE id = ?', (plan_id,)).fetchone()
    if not plan:
        return jsonify({'success': False, 'message': 'السجل غير موجود'}), 404
    pay_cols = {row['name'] for row in db.execute("PRAGMA table_info(payments)").fetchall()}
    if 'treatment_plan_id' in pay_cols:
        db.execute('DELETE FROM payments WHERE treatment_plan_id = ?', (plan_id,))
    db.execute('DELETE FROM treatment_plans WHERE id = ?', (plan_id,))
    db.commit()
    return jsonify({'success': True, 'message': 'تم حذف السجل بنجاح'})


@finance_bp.route('/installments')
@login_required
def installments():
    """صفحة الأقساط المخصصة"""
    db = get_db()
    _ensure_finance_schema(db)

    try:
        plans = db.execute('''
            SELECT tp.*, p.full_name as patient_name
            FROM treatment_plans tp
            JOIN patients p ON tp.patient_id = p.id
            ORDER BY tp.created_at DESC
        ''').fetchall()
    except sqlite3.OperationalError:
        # Retry once after forcing schema migration for legacy DB snapshots.
        _ensure_finance_schema(db)
        plans = db.execute('''
            SELECT tp.*, p.full_name as patient_name
            FROM treatment_plans tp
            JOIN patients p ON tp.patient_id = p.id
            ORDER BY tp.created_at DESC
        ''').fetchall()

    # إحصائيات الأقساط
    total_plans        = len(plans)
    active_plans       = sum(1 for p in plans if p['status'] == 'active')
    total_installments = total_plans
    total_remaining    = sum(
        p['total_cost'] - p['amount_paid'] for p in plans if p['status'] == 'active'
    )

    patients = db.execute('SELECT id, full_name FROM patients ORDER BY id').fetchall()

    return render_template('installments.html',
        plans=plans,
        total_plans=total_plans,
        active_plans=active_plans,
        total_installments=total_installments,
        total_remaining=total_remaining,
        patients=patients,
    )


OVERDUE_GRACE_DAYS = 20  # أيام التأخير المسموح بها بعد موعد القسط
OVERDUE_CYCLE_DAYS  = 30  # دورة القسط الشهرية (يوم)


def _calc_overdue(plans, today=None):
    """حساب المتأخرين بمنطق الأيام مع مهلة grace_days."""
    from datetime import date, timedelta
    if today is None:
        today = date.today()

    overdue_list = []
    for p in plans:
        created_date = date.fromisoformat(p['created_at'][:10])
        total_cost       = float(p['total_cost'])
        initial_payment  = float(p['initial_payment'] or 0)
        installments_cnt = int(p['installments_count'])
        amount_paid      = float(p['amount_paid'] or 0)

        if installments_cnt <= 0 or (total_cost - amount_paid) <= 0:
            continue

        monthly_inst = (total_cost - initial_payment) / installments_cnt
        if monthly_inst <= 0:
            continue

        # أحسب كم قسطاً استحق فعلاً (مع مهلة 3 أيام)
        installments_due = 0
        for n in range(1, installments_cnt + 1):
            due_date = created_date + timedelta(days=n * OVERDUE_CYCLE_DAYS)
            if today >= due_date + timedelta(days=OVERDUE_GRACE_DAYS):
                installments_due += 1

        if installments_due == 0:
            continue

        expected_paid  = initial_payment + installments_due * monthly_inst
        overdue_amount = expected_paid - amount_paid

        if overdue_amount < 0.5:
            continue

        # أقرب قسط قادم (للإشارة)
        next_due = None
        for n in range(installments_due + 1, installments_cnt + 2):
            nd = created_date + timedelta(days=n * OVERDUE_CYCLE_DAYS)
            if nd >= today:
                next_due = nd.isoformat()
                break

        # احسب عدد الأيام على آخر قسط مستحق
        last_due = created_date + timedelta(days=installments_due * OVERDUE_CYCLE_DAYS)
        days_late = (today - last_due).days

        overdue_list.append({
            'patient_name':   p['patient_name'],
            'patient_phone':  p['patient_phone'],
            'patient_id':     p['patient_id'],
            'plan_id':        p['id'],
            'description':    p['description'],
            'total_cost':     total_cost,
            'amount_paid':    amount_paid,
            'remaining':      total_cost - amount_paid,
            'expected_paid':  expected_paid,
            'overdue_amount': overdue_amount,
            'overdue_months': installments_due,
            'days_late':      days_late,
            'monthly_inst':   monthly_inst,
            'next_due':       next_due,
            'created_at':     p['created_at'][:10],
        })

    overdue_list.sort(key=lambda x: x['days_late'], reverse=True)
    return overdue_list


@finance_bp.route('/overdue')
@login_required
def overdue_patients():
    """قائمة المرضى المتأخرين عن سداد الأقساط (مع مهلة 3 أيام)"""
    db = get_db()
    plans = db.execute('''
        SELECT tp.*, p.full_name as patient_name, p.phone as patient_phone
        FROM treatment_plans tp
        JOIN patients p ON tp.patient_id = p.id
        WHERE tp.status = 'active' AND tp.installments_count > 0
          AND (tp.total_cost - tp.amount_paid) > 0
        ORDER BY tp.created_at ASC
    ''').fetchall()

    overdue_list  = _calc_overdue(plans)
    total_overdue = sum(r['overdue_amount'] for r in overdue_list)

    return render_template('overdue_patients.html',
        overdue_list=overdue_list,
        total_overdue=total_overdue,
    )


@finance_bp.route('/patient/<int:patient_id>/plans')
@login_required
def patient_plans(patient_id):
    """خطط العلاج الخاصة بمريض (JSON)"""
    db = get_db()
    plans = db.execute(
        'SELECT * FROM treatment_plans WHERE patient_id = ? AND status = "active"',
        (patient_id,)
    ).fetchall()
    return jsonify([dict(p) for p in plans])


# ─── تصدير Excel ──────────────────────────────────────────────────────

def _ensure_openpyxl():
    """محاولة تحميل openpyxl وقت التشغيل (يدعم التثبيت بعد تشغيل السيرفر)"""
    global openpyxl, Font, Alignment, PatternFill

    if openpyxl is not None and Font is not None and Alignment is not None and PatternFill is not None:
        return True

    try:
        openpyxl = importlib.import_module('openpyxl')
        styles = importlib.import_module('openpyxl.styles')
        Font = styles.Font
        Alignment = styles.Alignment
        PatternFill = styles.PatternFill
        return True
    except Exception:
        openpyxl = None
        Font = Alignment = PatternFill = None
        return False

def _excel_dependency_error():
    """رسالة موحدة عند عدم توفر مكتبة Excel"""
    return jsonify({
        'success': False,
        'message': 'ميزة Excel تحتاج تثبيت openpyxl داخل بيئة المشروع: pip install openpyxl. إذا تم التثبيت الآن فأعد تشغيل السيرفر.'
    }), 500

def _make_excel_style(ws, header_row, col_widths):
    """تنسيق ورقة Excel: عناوين + عرض الأعمدة + RTL"""
    ws.sheet_view.rightToLeft = True
    header_fill = PatternFill("solid", fgColor="4F46E5")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri")
    center = Alignment(horizontal="center", vertical="center")
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width


@finance_bp.route('/export/payments/excel')
@login_required
def export_payments_excel():
    """تصدير سجل المدفوعات إلى Excel"""
    if not _ensure_openpyxl():
        return _excel_dependency_error()

    db = get_db()
    payments = db.execute('''
        SELECT pay.*, p.full_name AS patient_name
        FROM payments pay
        JOIN patients p ON pay.patient_id = p.id
        ORDER BY pay.payment_date DESC
    ''').fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "سجل المدفوعات"

    headers = ['#', 'رقم المريض', 'اسم المريض', 'المبلغ (د.ع)', 'التاريخ', 'طريقة الدفع', 'ملاحظات']
    ws.append(headers)

    methods = {'cash': 'نقداً', 'card': 'بطاقة', 'transfer': 'حوالة'}
    for i, pay in enumerate(payments, 1):
        ws.append([
            i,
            pay['patient_id'],
            pay['patient_name'],
            float(pay['amount']),
            pay['payment_date'],
            methods.get(pay['payment_method'], pay['payment_method']),
            pay['notes'] or '',
        ])

    _make_excel_style(ws, 1, [5, 12, 22, 15, 14, 14, 25])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = 'attachment; filename="payments.xlsx"'
    return resp


@finance_bp.route('/export/plans/excel')
@login_required
def export_plans_excel():
    """تصدير خطط العلاج إلى Excel"""
    if not _ensure_openpyxl():
        return _excel_dependency_error()

    db = get_db()
    plans = db.execute('''
        SELECT tp.*, p.full_name AS patient_name
        FROM treatment_plans tp
        JOIN patients p ON tp.patient_id = p.id
        ORDER BY tp.created_at DESC
    ''').fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "خطط العلاج"

    headers = ['#', 'رقم المريض', 'اسم المريض', 'الوصف', 'التكلفة الكلية', 'المدفوع', 'المتبقي', 'الحالة', 'التاريخ']
    ws.append(headers)

    for i, plan in enumerate(plans, 1):
        remaining = float(plan['total_cost']) - float(plan['amount_paid'])
        ws.append([
            i,
            plan['patient_id'],
            plan['patient_name'],
            plan['description'],
            float(plan['total_cost']),
            float(plan['amount_paid']),
            remaining,
            'مكتملة' if plan['status'] == 'completed' else 'نشطة',
            plan['created_at'][:10] if plan['created_at'] else '',
        ])

    _make_excel_style(ws, 1, [5, 12, 22, 30, 16, 14, 14, 10, 14])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = 'attachment; filename="treatment_plans.xlsx"'
    return resp


# ─── فاتورة PDF (HTML للطباعة) ────────────────────────────────────────

@finance_bp.route('/invoice/<int:plan_id>')
@login_required
def invoice(plan_id):
    """عرض فاتورة خطة علاج قابلة للطباعة كـ PDF"""
    db = get_db()
    plan = db.execute('''
        SELECT tp.*, p.full_name AS patient_name, p.phone AS patient_phone
        FROM treatment_plans tp
        JOIN patients p ON tp.patient_id = p.id
        WHERE tp.id = ?
    ''', (plan_id,)).fetchone()

    if not plan:
        return "الفاتورة غير موجودة", 404

    payments = db.execute('''
        SELECT *,
               ROW_NUMBER() OVER (
                   PARTITION BY treatment_plan_id
                   ORDER BY payment_date, id
               ) as payment_num
        FROM payments
        WHERE treatment_plan_id = ?
        ORDER BY payment_date, id
    ''', (plan_id,)).fetchall()

    clinic = db.execute("SELECT key, value FROM settings WHERE key LIKE 'clinic_%'").fetchall()
    clinic_info = {row['key']: row['value'] for row in clinic}

    return render_template('invoice.html',
        plan=plan,
        payments=payments,
        clinic=clinic_info,
    )


@finance_bp.route('/receipt/<int:payment_id>')
@login_required
def receipt(payment_id):
    """إيصال دفعة واحدة قابل للطباعة"""
    db = get_db()
    payment = db.execute('''
        SELECT pay.*, p.full_name, p.phone
        FROM payments pay
        JOIN patients p ON pay.patient_id = p.id
        WHERE pay.id = ?
    ''', (payment_id,)).fetchone()

    if not payment:
        return "الإيصال غير موجود", 404

    clinic_rows = db.execute("SELECT key, value FROM settings WHERE key LIKE 'clinic_%'").fetchall()
    clinic = {r['key']: r['value'] for r in clinic_rows}

    return render_template('receipt.html', payment=payment, clinic=clinic)

